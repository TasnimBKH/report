#!/usr/bin/env python
# coding: utf-8

# In[4]:


import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver import ActionChains
import openpyxl
import os
import os.path


options = webdriver.ChromeOptions()
options.add_experimental_option('prefs', {"download.default_directory": "C:\\Users\\rafiquel.islam\\Desktop\\Report File",                                         #Change default directory for downloads
"download.prompt_for_download": False, #To auto download the file
"download.directory_upgrade": True,
"plugins.always_open_pdf_externally": True #It will not show PDF directly in chrome
})


driver = webdriver.Chrome(executable_path="E:\Software\chromedriver.exe",options=options)
driver.get("http://192.168.20.159/FinUltimusSPARKTEST/")
user_ele=driver.find_element_by_name("UserId")

pass_ele=driver.find_element_by_name("PasswordString")

user_ele.send_keys("s1")
pass_ele.send_keys("1")
pass_ele=driver.find_element_by_name("PasswordString")
driver.find_element_by_id("btnlogin").click()

drp=Select(driver.find_element_by_name("ctl00$ddlMenutype"))
#driver.find_element_by_value("R").click()
drp.select_by_value("R")


#v_search2=[9309,9302,9309,9830]
path="E:selenium//result.xlsx"
df=openpyxl.load_workbook(path)
sheet=df.active
v_search2=[]
df1=df['result']
df2=df['perameter']
for i in range(1,df1.max_row+1):
    j=df1.cell(row=i,column=1)
    v_search2.append(j.value)
print(v_search2)


for x in range(len(v_search2)):
    driver.find_element_by_id("txtMenuSearch").clear()
    ser_ele=driver.find_element_by_id("txtMenuSearch")
    k=v_search2[x]
    print(k)  
    ser_ele.send_keys(k)
    ser_ele.send_keys(Keys.ENTER)
    
    table=driver.find_elements_by_tag_name('h1')
    print(len(table))
    print("")
    if(len(table)==0):
        #print("welcome")
        for i in range(2,df2.max_row+1):
            
            c=df2.cell(row=i,column=1)
            #print(i,1)
            if c.value==k:
                
                #print('ok')
                d=df2.cell(row=i,column=4)
                id_tag=df2.cell(row=i,column=3)
                e=df2.cell(row=i,column=5)
                    
                if id_tag.value=='Id':
                    
                    #print("hello")
                    acc_ele=driver.find_element_by_id(d.value)
                    acc_ele.send_keys(e.value)
                elif id_tag.value=='name':
                    acc_ele=driver.find_element_by_name(d.value)
                    acc_ele.send_keys(e.value)
                        
                        
         
        ok_ele=driver.find_element_by_name("ctl00$contPlcHdrMasterHolder$btnOK").click()
     
        action = ActionChains(driver)
 
        #perform the operation
        action.key_down(Keys.ALT).send_keys('O').perform()
       
        #ok button work...............
       
        #print(driver.window_handles[0])
       
       
       
        window_after = driver.window_handles[1]
        #print(window_after)
        driver.switch_to.window(window_after)
        
        
        #print(window_after)
   
        suc_ele=driver.find_elements_by_tag_name('body')
        fal_ele=driver.find_elements_by_tag_name('form')
            

        print(k)
        for i in range(1,df1.max_row+1):
            for j in range(1,2):
               
                if(df1.cell(i,j).value==k ):
                 
                    if len(fal_ele)==0:
                    
                        df1.cell(i,j+1).value="ok"
                       
                    else:
                        
                        df1.cell(i,j+1).value="Failed"
      
        df.save(path)
        driver.close()
    
        file_exist=os.path.exists('C:\\Users\\rafiquel.islam\\Desktop\\Report File\\BU_ReportPreviewPOPUPUI.pdf')
        if(file_exist==True):
            os.rename("C:\\Users\\rafiquel.islam\\Desktop\\Report File\\BU_ReportPreviewPOPUPUI.pdf",
                        "C:\\Users\\rafiquel.islam\\Desktop\\Report File\\"+str(k)+".pdf")
            
        print(file_exist)
       
        window_before = driver.window_handles[0]
        driver.switch_to.window(window_before)
             
             
    else:
        for i in range(1,df1.max_row+1):
            for j in range(1,2):
               
                if(df1.cell(i,j).value==k ):
                     df1.cell(i,j+1).value="Unauthorized"
        df.save(path)       
                    
        continue


# # Give Valid variable Name

# In[ ]:


import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver import ActionChains
import openpyxl
import os
import os.path


options = webdriver.ChromeOptions()
options.add_experimental_option('prefs', {"download.default_directory": "C:\\Users\\rafiquel.islam\\Desktop\\Report File",                                         #Change default directory for downloads
"download.prompt_for_download": False, #To auto download the file
"download.directory_upgrade": True,
"plugins.always_open_pdf_externally": True #It will not show PDF directly in chrome
})


driver = webdriver.Chrome(executable_path="E:\Software\chromedriver.exe",options=options)
driver.get("http://192.168.20.159/FinUltimusSPARKTEST/")
user_ele=driver.find_element_by_name("UserId")

pass_ele=driver.find_element_by_name("PasswordString")

user_ele.send_keys("s1")
pass_ele.send_keys("1")
pass_ele=driver.find_element_by_name("PasswordString")
driver.find_element_by_id("btnlogin").click()

drp=Select(driver.find_element_by_name("ctl00$ddlMenutype"))
#driver.find_element_by_value("R").click()
drp.select_by_value("R")


#v_search2=[9309,9302,9309,9830]
path="E:selenium//result.xlsx"
df=openpyxl.load_workbook(path)
sheet=df.active
v_search2=[]
df1=df['result']
df2=df['perameter']
for i in range(1,df1.max_row+1):
    j=df1.cell(row=i,column=1)
    v_search2.append(j.value)
print(v_search2)


for x in range(len(v_search2)):
    driver.find_element_by_id("txtMenuSearch").clear()
    ser_ele=driver.find_element_by_id("txtMenuSearch")
    k=v_search2[x]
    print(k)  
    ser_ele.send_keys(k)
    ser_ele.send_keys(Keys.ENTER)
    
    table=driver.find_elements_by_tag_name('h1')
    print(len(table))
    print("")
    if(len(table)==0):
        #print("welcome")
        for i in range(2,df2.max_row+1):
            
            c=df2.cell(row=i,column=1)
            #print(i,1)
            if c.value==k:
                
                #print('ok')
                d=df2.cell(row=i,column=4)
                id_tag=df2.cell(row=i,column=3)
                e=df2.cell(row=i,column=5)
                    
                if id_tag.value=='Id':
                    
                    #print("hello")
                    acc_ele=driver.find_element_by_id(d.value)
                    acc_ele.send_keys(e.value)
                elif id_tag.value=='name':
                    acc_ele=driver.find_element_by_name(d.value)
                    acc_ele.send_keys(e.value)
                        
                        
         
        ok_ele=driver.find_element_by_name("ctl00$contPlcHdrMasterHolder$btnOK").click()
     
        action = ActionChains(driver)
 
        #perform the operation
        action.key_down(Keys.ALT).send_keys('O').perform()
       
        #ok button work...............
       
        #print(driver.window_handles[0])
       
       
       
        window_after = driver.window_handles[1]
        #print(window_after)
        driver.switch_to.window(window_after)
        
        
        #print(window_after)
   
        suc_ele=driver.find_elements_by_tag_name('body')
        fal_ele=driver.find_elements_by_tag_name('form')
            

        print(k)
        for i in range(1,df1.max_row+1):
            for j in range(1,2):
               
                if(df1.cell(i,j).value==k ):
                 
                    if len(fal_ele)==0:
                    
                        df1.cell(i,j+1).value="ok"
                       
                    else:
                        
                        df1.cell(i,j+1).value="Failed"
      
        df.save(path)
        driver.close()
    
        file_exist=os.path.exists('C:\\Users\\rafiquel.islam\\Desktop\\Report File\\BU_ReportPreviewPOPUPUI.pdf')
        if(file_exist==True):
            os.rename("C:\\Users\\rafiquel.islam\\Desktop\\Report File\\BU_ReportPreviewPOPUPUI.pdf",
                        "C:\\Users\\rafiquel.islam\\Desktop\\Report File\\"+str(k)+".pdf")
            
        print(file_exist)
       
        window_before = driver.window_handles[0]
        driver.switch_to.window(window_before)
             
             
    else:
        for i in range(1,df1.max_row+1):
            for j in range(1,2):
               
                if(df1.cell(i,j).value==k ):
                     df1.cell(i,j+1).value="Unauthorized"
        df.save(path)       
                    
        continue


# # Create function 

# In[7]:


import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver import ActionChains
import openpyxl
import os
import os.path


options = webdriver.ChromeOptions()
options.add_experimental_option('prefs', {"download.default_directory": "C:\\Users\\rafiquel.islam\\Desktop\\Report File",                                         #Change default directory for downloads
"download.prompt_for_download": False, #To auto download the file
"download.directory_upgrade": True,
"plugins.always_open_pdf_externally": True #It will not show PDF directly in chrome
})


driver = webdriver.Chrome(executable_path="E:\Software\chromedriver.exe",options=options)
driver.get("http://192.168.20.159/FinUltimusSPARKTEST/")
user_ele=driver.find_element_by_name("UserId")

pass_ele=driver.find_element_by_name("PasswordString")

user_ele.send_keys("s1")
pass_ele.send_keys("1")
pass_ele=driver.find_element_by_name("PasswordString")
driver.find_element_by_id("btnlogin").click()

drp=Select(driver.find_element_by_name("ctl00$ddlMenutype"))
#driver.find_element_by_value("R").click()
drp.select_by_value("R")


#v_search2=[9309,9302,9309,9830]
path="E:selenium//result.xlsx"
df=openpyxl.load_workbook(path)
sheet=df.active
v_search2=[]
df1=df['result']
df2=df['perameter']
for i in range(1,df1.max_row+1):
    j=df1.cell(row=i,column=1)
    v_search2.append(j.value)
print(v_search2)


for x in range(len(v_search2)):
    driver.find_element_by_id("txtMenuSearch").clear()
    ser_ele=driver.find_element_by_id("txtMenuSearch")
    k=v_search2[x]
    print(k)  
    ser_ele.send_keys(k)
    ser_ele.send_keys(Keys.ENTER)
    
    table=driver.find_elements_by_tag_name('h1')
    print(len(table))
    print("")
    if(len(table)==0):
        #print("welcome")
        for i in range(2,df2.max_row+1):
            
            c=df2.cell(row=i,column=1)
            #print(i,1)
            if c.value==k:
                
                #print('ok')
                d=df2.cell(row=i,column=4)
                id_tag=df2.cell(row=i,column=3)
                e=df2.cell(row=i,column=5)
                    
                if id_tag.value=='Id':
                    
                    #print("hello")
                    acc_ele=driver.find_element_by_id(d.value)
                    acc_ele.send_keys(e.value)
                elif id_tag.value=='name':
                    acc_ele=driver.find_element_by_name(d.value)
                    acc_ele.send_keys(e.value)
                        
                        
         
        ok_ele=driver.find_element_by_name("ctl00$contPlcHdrMasterHolder$btnOK").click()
     
        action = ActionChains(driver)
 
        #perform the operation
        action.key_down(Keys.ALT).send_keys('O').perform()
       
        #ok button work...............
       
        #print(driver.window_handles[0])
       
       
       
        window_after = driver.window_handles[1]
        #print(window_after)
        driver.switch_to.window(window_after)
        
        
        #print(window_after)
   
        suc_ele=driver.find_elements_by_tag_name('body')
        fal_ele=driver.find_elements_by_tag_name('form')
            

        print(k)
        for i in range(1,df1.max_row+1):
            for j in range(1,2):
               
                if(df1.cell(i,j).value==k ):
                 
                    if len(fal_ele)==0:
                    
                        df1.cell(i,j+1).value="ok"
                       
                    else:
                        
                        df1.cell(i,j+1).value="Failed"
      
        df.save(path)
        driver.close()
    
        file_exist=os.path.exists('C:\\Users\\rafiquel.islam\\Desktop\\Report File\\BU_ReportPreviewPOPUPUI.pdf')
        if(file_exist==True):
            os.rename("C:\\Users\\rafiquel.islam\\Desktop\\Report File\\BU_ReportPreviewPOPUPUI.pdf",
                        "C:\\Users\\rafiquel.islam\\Desktop\\Report File\\"+str(k)+".pdf")
        time.sleep(2)   
        print(file_exist)
       
        window_before = driver.window_handles[0]
        driver.switch_to.window(window_before)
             
             
    else:
        for i in range(1,df1.max_row+1):
            for j in range(1,2):
               
                if(df1.cell(i,j).value==k ):
                     df1.cell(i,j+1).value="Unauthorized"
        df.save(path)       
                    
        continue


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




