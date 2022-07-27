#!/usr/bin/env python
# coding: utf-8

# In[2]:


import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver import ActionChains
import openpyxl
import os
import os.path
import asyncio


df=openpyxl.load_workbook("E:selenium//result.xlsx")
sheet=df.active
fastPathSearch=[]
resultsheet=df['result']
perametersheet=df['perameter']
document=df['document']
website_path=document.cell(1,2).value
driver_link=document.cell(2,2).value
user=document.cell(3,2).value
password=document.cell(4,2).value
input_output_path=document.cell(5,2).value
pdf_folder_path=document.cell(6,2).value


options = webdriver.ChromeOptions()
options.add_experimental_option('prefs', {"download.default_directory":pdf_folder_path ,  #Change default directory for downloads
"download.prompt_for_download": False, #To auto download the file
"download.directory_upgrade": True,
"plugins.always_open_pdf_externally": True #It will not show PDF directly in chrome
})

driver = webdriver.Chrome(executable_path=driver_link,options=options)
driver.get(website_path)


def main():
    
    login(user,password)
    input_list()
    
    for x in range(len( fastPathSearch)):
        driver.find_element_by_id("txtMenuSearch").clear()
        ser_ele=driver.find_element_by_id("txtMenuSearch")
        path_id= fastPathSearch[x]
        ser_ele.send_keys(path_id)
        ser_ele.send_keys(Keys.ENTER)
        check_unauthorize=driver.find_elements_by_tag_name('h1')
        print(len(check_unauthorize))
        if(len(check_unauthorize)==0):
          
            for i in range(2,perametersheet.max_row+1):

                c=perametersheet.cell(row=i,column=1)
               
                if c.value==path_id:
                    d=perametersheet.cell(row=i,column=4)
                    id_tag=perametersheet.cell(row=i,column=3)
                    e=perametersheet.cell(row=i,column=5)
                    if id_tag.value=='Id':

                        acc_ele=driver.find_element_by_id(d.value)
                        acc_ele.send_keys(e.value)
                    elif id_tag.value=='name':
                        acc_ele=driver.find_element_by_name(d.value)
                        acc_ele.send_keys(e.value)


         
            ok_ele=driver.find_element_by_name("ctl00$contPlcHdrMasterHolder$btnOK").click()

            action = ActionChains(driver)
            action.key_down(Keys.ALT).send_keys('O').perform()

            window_after = driver.window_handles[1]
            driver.switch_to.window(window_after)
            
            fal_ele=driver.find_elements_by_tag_name('form')
            for i in range(1,resultsheet.max_row+1):
                for j in range(1,2):

                    if(resultsheet.cell(i,j).value==path_id ):

                        if len(fal_ele)==0:

                            resultsheet.cell(i,j+1).value="ok"

                        else:

                            resultsheet.cell(i,j+1).value="Failed"

            df.save(input_output_path)


            file_exist=os.path.exists('C:\\Users\\rafiquel.islam\\Desktop\\Report File\\BU_ReportPreviewPOPUPUI.pdf')
            #await asyncio.sleep(5) 
            if(file_exist==True):
                os.rename("C:\\Users\\rafiquel.islam\\Desktop\\Report File\\BU_ReportPreviewPOPUPUI.pdf",
                          "C:\\Users\\rafiquel.islam\\Desktop\\Report File\\"+str(path_id)+".pdf")
                #await asyncio.sleep(3)
            #time.sleep(0.5)   
           
            driver.close()

            window_before = driver.window_handles[0]
            driver.switch_to.window(window_before)
        
        else:
            for i in range(1,resultsheet.max_row+1):
                for j in range(1,2):

                    if(resultsheet.cell(i,j).value== path_id ):
                         resultsheet.cell(i,j+1).value="Unauthorized"
            df.save(input_output_path)       

            continue

            
            
            
            
            
            
        
def login(username,password):
   
    user_ele=driver.find_element_by_name("UserId")

    pass_ele=driver.find_element_by_name("PasswordString")

    user_ele.send_keys(username)
    pass_ele.send_keys(password)
    driver.find_element_by_id("btnlogin").click()
    drp=Select(driver.find_element_by_name("ctl00$ddlMenutype"))
    drp.select_by_value("R")
    
   
    
def input_list():
    #df=openpyxl.load_workbook(input_output_path)
    #sheet=df.active
    #resultsheet=df['result']
    #perametersheet=df['perameter']
    for i in range(1,resultsheet.max_row+1):
        j=resultsheet.cell(row=i,column=1)
        fastPathSearch.append(j.value)
   
    
def search_path():
    for x in range(len( fastPathSearch)):
        driver.find_element_by_id("txtMenuSearch").clear()
        ser_ele=driver.find_element_by_id("txtMenuSearch")
        k= fastPathSearch[x]
        print(k)  
        ser_ele.send_keys(k)
        ser_ele.send_keys(Keys.ENTER)
        check_unauthorize=driver.find_elements_by_tag_name('h1')
        print(len(check_unauthorize))
        return  check_unauthorize
   
    



main()


# In[ ]:




